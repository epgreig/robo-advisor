import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.chart import (
    PieChart,
    BarChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.styles import colors
from openpyxl.cell import Cell

wb = xl.Workbook()
wb.remove(wb.active)

ws_comp = wb.create_sheet("Composition", 0)
ws_comp.sheet_properties.tabColor = "DC143C"

ws_perf = wb.create_sheet("Performance", 2)
ws_perf.sheet_properties.tabColor = "228B22"

ws_risk = wb.create_sheet("Risk Analytics", 1)
ws_risk.sheet_properties.tabColor = "FFD700"

sheets = [ws_comp, ws_perf, ws_risk]

# Three worksheets for each type of analytic:
#
# Composition
# Backward Peformance Analytics
# Forward Risk Analytics

etfs = ['SPY', 'EFA', 'XLF', 'XLK', 'XLV', 'XLP', 'XLE', 'EWJ', 'EWZ', 'XLU', 'XLI', 'EZU', 'IYR', 'XLB',
        'RWR', 'IXN', 'ISMUF', 'ICF', 'IYZ', 'ILF', 'IEV', 'TIP', 'AGG', 'IEF', 'TLT', 'SHY', 'LQD']

types = ['EQ', 'FI', 'EM', 'RE', 'OPT']

risks = ['Delta', 'Gamma', 'Vega', 'Theta', 'Rho']

darkBlue  = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
lightBlue = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')

for array in [ws_comp['AJ1':'AK1'],
              ws_comp['AJ2':'AJ6']]:
    for rng in array:
        for cell in rng:
            cell.fill = darkBlue

for array in [ws_comp['AK2':'AK6']]:
    for rng in array:
        for cell in rng:
            cell.fill = lightBlue

for row, risk in zip(ws_comp.iter_rows(min_row=2, min_col=36, max_col=36, max_row=len(risks)+1), risks):
    for cell in row:
        cell.value = risk

for ws in sheets:
    for array in [ws['A1':'B1'],
                 ws['A2':'A6'],
                 ws['K1':'L1'],
                 ws['K2':'K28']]:
        for rng in array:
            for cell in rng:
                cell.fill = darkBlue

    for array in [ws['B2':'B6'],
                  ws['L2':'L28']]:
        for rng in array:
            for cell in rng:
                cell.fill = lightBlue

    for row, aClass in zip(ws.iter_rows(min_row=2, min_col=1, max_col=1, max_row=len(types)+1), types):
        for cell in row:
            cell.value = aClass

    for row, etf in zip(ws.iter_rows(min_row=2, min_col=11, max_col=11, max_row=len(etfs)+1), etfs):
        for cell in row:
            cell.value = etf

ws_comp['B1'].value = 'Weight by Asset Class'
ws_comp['L1'].value = 'Weight by ETF'
ws_comp['AK1'].value = 'Portfolio Risk Value'

######################
# PLACEHOLDER VALUES #
##################################
for r in range(2, 2+len(types)):
    ws_comp.cell(row=r, column=2).value = 1./float(len(types))
for r in range(2, 2+len(etfs)):
    ws_comp.cell(row=r, column=12).value = 1./float(len(etfs))
sign = 1
for r in range(2, 2+len(risks)):
    sign *= -1
    ws_comp.cell(row=r, column=37).value = sign * 2.
##################################

def make_pie(sheet, left_col, top_row, bot_row, title, print_cell, height, width):
    left_col = int(left_col)
    right_col = left_col + 1
    top_row = int(top_row)
    bot_row = int(bot_row)
    title = str(title)
    print_cell = str(print_cell)
    height = float(height)
    width = float(width)

    pie = PieChart()
    labels = Reference(sheet, min_col=left_col, max_col=left_col, min_row=top_row+1, max_row=bot_row)
    data = Reference(sheet, min_col=right_col, max_col=right_col, min_row=top_row, max_row=bot_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = title
    pie.height = height
    pie.width = width
    sheet.add_chart(pie, print_cell)

    return None

def make_bar(sheet, left_col, top_row, bot_row, title, x_title, y_title, print_cell, height, width):
    left_col = int(left_col)
    right_col = left_col + 1
    top_row = int(top_row)
    bot_row = int(bot_row)
    title = str(title)
    print_cell = str(print_cell)
    height = float(height)
    width = float(width)

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.shape = 4
    bar.title = title
    bar.y_axis.title = y_title
    bar.x_axis.title = x_title
    labels = Reference(sheet, min_col=left_col, max_col=left_col, min_row=top_row+1, max_row=bot_row)
    data = Reference(sheet, min_col=right_col, max_col=right_col, min_row=top_row, max_row=bot_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.height = height
    bar.width = width

    sheet.add_chart(bar, print_cell)

make_pie(ws_comp, 1, 1, len(types), 'Weight by Asset Class', 'A8', 7.5, 8.5)
make_pie(ws_comp, 11, 1, 28, 'Weight by ETF', 'M8', 7.5, 8.5)
make_bar(ws_comp, 36, 1, len(risks), 'Portfolio Option Risk', 'Risks', 'Value', 'AJ8', 7.5, 12)

for ws in sheets:
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.coordinate in ws.merged_cells:
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

wb.save("Report.xlsx")
